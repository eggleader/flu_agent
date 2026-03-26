"""
创建 BioAgent 技术栈对比表格
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
sheet = wb.active
sheet.title = "技术方案对比"

# 表头样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
bioagent_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
original_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 表头
headers = ["技术类别", "原方案设计", "POPGENAGENT", "BioAgent 实际采用", "不同技术优劣比较"]
for col, header in enumerate(headers, 1):
    cell = sheet.cell(1, col, header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# 数据行
data = [
    ["后端框架", "Python FastAPI\n(异步性能好，自动生成API文档)", "Python Flask/自研\n(轻量框架)", "Python 纯实现\n(零框架依赖，简洁轻量)", "FastAPI：功能完善但有学习曲线\nFlask：灵活但需自行组装\n纯Python：代码量增加但无外部依赖"],
    ["前端框架", "React + Ant Design\n(组件丰富)", "需自行开发\n(无内置UI)", "Gradio\n(Web界面，快速开发)", "React+AntD：专业但开发周期长\n需开发：完全定制但从零开始\nGradio：5分钟完成原型，够用即可"],
    ["LLM 引擎", "LangChain + 本地LLM\n或 OpenAI API", "未指定\n(Planner+Executor架构)", "Ollama\n(本地推理)", "LangChain：封装完善但抽象过度\n未指定：灵活性高但实现成本大\nOllama：本地零成本，隐私好"],
    ["工具协议", "LangChain Agents", "未指定\n(Planner调用工具)", "OpenAI Function Calling", "LangChain：标准化但受框架限制\nFC：轻量标准，兼容性更好"],
    ["数据库", "PostgreSQL + Neo4j\n(关系+图)", "无\n(无持久化)", "SQLite\n(对话历史持久化)", "PG+Neo4j：功能强大但部署复杂\n无：无状态，每次重启重置\nSQLite：轻量零运维，单文件易备份"],
    ["工作流引擎", "Nextflow\n(专业工作流)", "无", "YAML 驱动\n(自定义引擎)", "Nextflow：专业但学习曲线陡\n无：简单但缺乏流程管理\nYAML：灵活简单，但功能有限"],
    ["任务队列", "Celery + Redis\n(异步任务)", "无", "无\n(同步执行)", "Celery+Redis：解耦但增加复杂度\n无：简单可靠，但长任务阻塞"],
    ["容器化", "Docker", "Docker", "Docker", "一致：环境一致性保障"],
    ["报告生成", "Jinja2 + WeasyPrint\n+ Plotly", "Plotly\n(可视化)", "Plotly + Jinja2", "完整PDF vs 可视化Web展示"],
    ["文件存储", "本地 + MinIO\n(可扩展)", "本地文件系统", "本地文件系统", "MinIO分布式 vs 本地简单"],
    ["会话管理", "无/需自行实现", "无", "SQLite\n(会话历史)", "有会话：上下文连续\n无：每次独立对话"],
    ["版本控制", "Git + GitHub/GitLab", "Git + GitHub", "Git + GitHub", "一致"],
    ["监控日志", "Prometheus + Grafana\n+ ELK", "无", "无", "完整监控 vs 轻量无监控"],
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = sheet.cell(row_idx, col_idx, value)
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = thin_border
        
        if col_idx == 4:  # BioAgent 列
            cell.fill = bioagent_fill
        elif col_idx == 2:  # 原方案列
            cell.fill = original_fill

# 设置列宽
sheet.column_dimensions['A'].width = 12
sheet.column_dimensions['B'].width = 25
sheet.column_dimensions['C'].width = 20
sheet.column_dimensions['D'].width = 20
sheet.column_dimensions['E'].width = 38

# 设置行高
for row in range(1, len(data) + 2):
    sheet.row_dimensions[row].height = 45

# 添加总结 sheet
summary_sheet = wb.create_sheet("架构差异总结")
summary_sheet['A1'] = "BioAgent vs 原方案 架构差异总结"
summary_sheet['A1'].font = Font(bold=True, size=14)
summary_sheet.merge_cells('A1:C1')

summary_data = [
    ["", "说明"],
    ["设计理念", "原方案面向产品化生产环境，强调高可用、可扩展；BioAgent 面向原型验证，强调轻量、快速"],
    ["复杂度", "原方案 11 个技术组件；BioAgent 6 个核心组件（去框架化）"],
    ["扩展性", "原方案预留 PostgreSQL/Neo4j/MinIO/Prometheus 等扩展点；BioAgent 当前聚焦核心功能"],
    ["适用场景", "原方案：大规模团队协作、产品级部署；BioAgent：个人/小团队研究、原型开发"],
]

for row_idx, row_data in enumerate(summary_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = summary_sheet.cell(row_idx, col_idx, value)
        if row_idx == 3:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")

summary_sheet.column_dimensions['A'].width = 15
summary_sheet.column_dimensions['B'].width = 50

wb.save('/Users/guojd/CodeBuddy/Claw/projects/agent/技术方案对比.xlsx')
print("Excel 表格已生成: /Users/guojd/CodeBuddy/Claw/projects/agent/技术方案对比.xlsx")
